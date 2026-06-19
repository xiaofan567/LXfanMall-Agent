"""
deepdoc/excel_parser.py - Excel 文档解析器

使用 openpyxl 读取 Excel，将每行转换为自然语言描述。
"""

from __future__ import annotations

import io
import logging
from pathlib import Path

from .base import BaseParser, DocumentChunk, ParserRegistry
from .utils import clean_text

logger = logging.getLogger(__name__)


@ParserRegistry.register(["xlsx", "xls"])
class ExcelParser(BaseParser):
    """Excel 文档解析器。"""

    def parse(self, file_path: str | Path) -> list[DocumentChunk]:
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")

        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        try:
            wb = load_workbook(str(file_path), read_only=True, data_only=True)
            return self._extract_chunks(wb, file_path.name)
        except Exception as e:
            logger.error("Excel 解析失败: %s - %s", file_path, e)
            raise

    def parse_bytes(self, data: bytes, filename: str = "") -> list[DocumentChunk]:
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        try:
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            return self._extract_chunks(wb, filename)
        except Exception as e:
            logger.error("Excel 字节数据解析失败: %s", e)
            raise

    def _extract_chunks(self, wb, source: str) -> list[DocumentChunk]:
        """从 Workbook 中提取 chunks。"""
        chunks: list[DocumentChunk] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # 第一行为表头
            headers = [str(h).strip() if h is not None else f"列{i+1}" for i, h in enumerate(rows[0])]
            data_rows = rows[1:]

            for row_idx, row in enumerate(data_rows, 2):
                # 跳过全空行
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                # 构建自然语言描述
                parts: list[str] = []
                for i, cell in enumerate(row):
                    if i >= len(headers):
                        break
                    value = self._format_cell(cell)
                    if value:  # 跳过空单元格
                        parts.append(f"{headers[i]}: {value}")

                if parts:
                    content = ", ".join(parts)
                    chunks.append(DocumentChunk(
                        content=content,
                        metadata={
                            "source": source,
                            "sheet": sheet_name,
                            "row": row_idx,
                            "chunk_type": "table_row",
                        },
                    ))

            # 同时生成一个完整的表格摘要 chunk（表头 + 前几行预览）
            preview_rows = data_rows[:5]
            if preview_rows:
                preview_lines = [f"工作表: {sheet_name}"]
                preview_lines.append(f"列: {', '.join(headers)}")
                preview_lines.append(f"共 {len(data_rows)} 行数据")
                for row in preview_rows:
                    parts = []
                    for i, cell in enumerate(row):
                        if i >= len(headers):
                            break
                        value = self._format_cell(cell)
                        if value:
                            parts.append(f"{headers[i]}={value}")
                    if parts:
                        preview_lines.append(" | ".join(parts))
                if len(data_rows) > 5:
                    preview_lines.append(f"... 共 {len(data_rows)} 行")

                chunks.insert(
                    # 插入到该 sheet 数据之前
                    max(0, len(chunks) - len(data_rows)),
                    DocumentChunk(
                        content="\n".join(preview_lines),
                        metadata={
                            "source": source,
                            "sheet": sheet_name,
                            "chunk_type": "table_summary",
                        },
                    ),
                )

        wb.close()
        logger.info("Excel 解析完成: %s (%d 个块)", source, len(chunks))
        return chunks

    @staticmethod
    def _format_cell(cell) -> str:
        """格式化单元格值为字符串。"""
        if cell is None:
            return ""
        if isinstance(cell, float):
            # 去除不必要的小数点: 100.0 -> 100
            if cell == int(cell):
                return str(int(cell))
            return f"{cell:.2f}".rstrip("0").rstrip(".")
        return str(cell).strip()
